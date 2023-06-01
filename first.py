import datetime
from openpyxl import Workbook
from openpyxl.styles import Alignment

attendance_times = {}

while True:
    name = input("Введите свое имя (или 'выход' для завершения): ")
    if name == 'выход':
        break

    if name not in attendance_times:
        # Фиксируем время прихода при первом вводе имени сотрудника
        arrival_time = datetime.datetime.now()
        attendance_times[name] = {'приход': arrival_time}
        print("Время прихода успешно записано.")
    else:
        action = input("Введите 'уход' для фиксации ухода: ")
        if action.lower() == 'уход':
            # Фиксируем время ухода при вводе имени и слова "уход"
            departure_time = datetime.datetime.now()
            attendance_times[name]['уход'] = departure_time
            print("Время ухода успешно записано.")
        else:
            print("Некорректное действие. Попробуйте еще раз.")

# Создаем новую рабочую книгу Excel
wb = Workbook()
# Выбираем активный лист
sheet = wb.active

# Записываем заголовки столбцов
sheet['A1'] = 'Имя'
sheet['B1'] = 'Время прихода'
sheet['C1'] = 'Время ухода'
sheet['D1'] = 'Количество отработанного времени'

# Записываем данные о приходе и уходе сотрудников
row = 2
for name, times in attendance_times.items():
    sheet.cell(row=row, column=1, value=name)
    sheet.cell(row=row, column=2, value=times['приход'])
    if 'уход' in times:
        sheet.cell(row=row, column=3, value=times['уход'])
        work_time = times['уход'] - times['приход']
        sheet.cell(row=row, column=4, value=str(work_time))
    else:
        sheet.cell(row=row, column=3, value="Сотрудник еще не ушел")
        sheet.cell(row=row, column=4, value="---")
    row += 1

# Выравниваем содержимое ячеек по центру
for row in sheet.iter_rows():
    for cell in row:
        cell.alignment = Alignment(horizontal='center')

# Автонастраиваем ширину столбцов
for column_cells in sheet.columns:
    max_length = 0
    column = column_cells[0].column_letter  # получаем буквенное представление столбца
    for cell in column_cells:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except TypeError:
            pass
    adjusted_width = (max_length + 2) * 1.2
    sheet.column_dimensions[column].width = adjusted_width

# Сохраняем файл Excel
wb.save('отчет.xlsx')
print("Отчет успешно сохранен в файле 'отчет.xlsx'")